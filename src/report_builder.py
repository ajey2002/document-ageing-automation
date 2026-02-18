from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .excel_styles import setup_table_styles
from .config import (
    COL_COMPANY,
    COL_DOC_DATE,
    COL_DOC_TYPE,
    COL_DOC_CURR,
    COL_AMT_DOC,
    COL_LOC_CURR,
    COL_AMT_LOC,
    COL_TEXT,
)


def autosize_columns(ws, min_width=10, max_width=70):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def _apply_row_style(ws, row_idx: int, col_from: int, col_to: int, border, alignment):
    for c in range(col_from, col_to + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = alignment


def build_summary_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    styles = setup_table_styles()
    ws = wb.create_sheet(sheet_name)

    # Title row (Row 1) + keep Row 2 blank 
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Document Ageing Report as at {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font = styles["title_font"]
    ws["A1"].alignment = styles["title_align"]

    # Header row (Row 3)
    headers = [
        "Comapany",
        "Account",
        "Document currency",
        "Amount in doc. curr.",
        "Local Currency",
        "Amount in local currency",
    ]
    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["header_align"]
        cell.border = styles["border"]

    # Summary aggregation
    summary_df = (
        df.groupby([COL_COMPANY, "_AccountClean", COL_DOC_CURR, COL_LOC_CURR], dropna=False)
          .agg({COL_AMT_DOC: "sum", COL_AMT_LOC: "sum"})
          .reset_index()
          .sort_values(["_AccountClean", COL_COMPANY])
    )

    start_row = header_row + 1
    r = start_row
    for i in range(len(summary_df)):
        ws.cell(r, 1).value = summary_df.iloc[i][COL_COMPANY]
        ws.cell(r, 2).value = summary_df.iloc[i]["_AccountClean"]
        ws.cell(r, 3).value = summary_df.iloc[i][COL_DOC_CURR]
        ws.cell(r, 4).value = float(summary_df.iloc[i][COL_AMT_DOC])
        ws.cell(r, 5).value = summary_df.iloc[i][COL_LOC_CURR]
        ws.cell(r, 6).value = float(summary_df.iloc[i][COL_AMT_LOC])

        # Borders + vertical bottom alignment for entire row
        for c in range(1, 7):
            ws.cell(r, c).border = styles["border"]
            ws.cell(r, c).alignment = styles["bottom_center"]  # default for summary

        # Amount columns right aligned (D and F)
        ws.cell(r, 4).alignment = styles["bottom_right"]
        ws.cell(r, 6).alignment = styles["bottom_right"]

        r += 1

    # number formats for amount columns
    for rr in range(start_row, r):
        ws.cell(rr, 4).number_format = "#,##0.00;(#,##0.00)"
        ws.cell(rr, 6).number_format = "#,##0.00;(#,##0.00)"

    ws.freeze_panes = "A4"
    autosize_columns(ws)
    return ws


def build_account_sheet(wb: Workbook, account: str, sub_df: pd.DataFrame):
    styles = setup_table_styles()
    ws = wb.create_sheet(title=account[:31])

    # Keep Row 1 blank
    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Account: {account}  |  As at {date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = styles["title_font"]
    ws["A2"].alignment = styles["title_align"]

    headers = [
        "Comapany", "Account", "Document Date", "Document Type", "Document currency",
        "Amount in doc. curr.", "Local Currency", "Amount in local currency", "Text", "Doc Ageing"
    ]

    # Header row (Row 3)
    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["header_align"]
        cell.border = styles["border"]

    start_row = header_row + 1

    for i in range(len(sub_df)):
        r = start_row + i

        ws.cell(r, 1).value = sub_df.iloc[i][COL_COMPANY]
        ws.cell(r, 2).value = sub_df.iloc[i]["_AccountClean"]
        ws.cell(r, 3).value = sub_df.iloc[i][COL_DOC_DATE]
        ws.cell(r, 4).value = sub_df.iloc[i][COL_DOC_TYPE]
        ws.cell(r, 5).value = sub_df.iloc[i][COL_DOC_CURR]
        ws.cell(r, 6).value = sub_df.iloc[i][COL_AMT_DOC]
        ws.cell(r, 7).value = sub_df.iloc[i][COL_LOC_CURR]
        ws.cell(r, 8).value = sub_df.iloc[i][COL_AMT_LOC]
        ws.cell(r, 9).value = sub_df.iloc[i][COL_TEXT]

        # Doc Ageing (TODAY - Document Date col C)
        ws.cell(r, 10).value = f'=IF(C{r}="","",TODAY()-C{r})'

        # Apply borders + bottom alignment to all cells
        for c in range(1, 11):
            ws.cell(r, c).border = styles["border"]
            ws.cell(r, c).alignment = styles["bottom_left"]  # base

        # Alignments per requirement:
        # Center: Document Type (D), Document currency (E), Local Currency (G), Doc Ageing (J)
        ws.cell(r, 4).alignment = styles["bottom_center"]
        ws.cell(r, 5).alignment = styles["bottom_center"]
        ws.cell(r, 7).alignment = styles["bottom_center"]
        ws.cell(r, 10).alignment = styles["bottom_center"]

        # Center: Account + Document Date
        ws.cell(r, 2).alignment = styles["bottom_center"]
        ws.cell(r, 3).alignment = styles["bottom_center"]

        # Right: amounts
        ws.cell(r, 6).alignment = styles["bottom_right"]
        ws.cell(r, 8).alignment = styles["bottom_right"]

    last_data_row = start_row + len(sub_df) - 1
    total_row = last_data_row + 1 if len(sub_df) else start_row

    # Totals in F and H 
    if len(sub_df):
        ws.cell(total_row, 6).value = f"=SUM(F{start_row}:F{last_data_row})"
        ws.cell(total_row, 8).value = f"=SUM(H{start_row}:H{last_data_row})"

    # Totals styling: bold + double top border on the whole row
    for c in range(1, 11):
        cell = ws.cell(total_row, c)
        cell.border = styles["total_top_border"]
        cell.font = styles["total_font"]
        cell.alignment = styles["bottom_left"]

    # Alignment on totals row
    ws.cell(total_row, 6).alignment = styles["bottom_right"]
    ws.cell(total_row, 8).alignment = styles["bottom_right"]

    # Formats
    ws.freeze_panes = "A4"
    ws.column_dimensions["I"].width = 60

    for rr in range(start_row, total_row + 1):
        ws.cell(rr, 6).number_format = "#,##0.00;(#,##0.00)"
        ws.cell(rr, 8).number_format = "#,##0.00;(#,##0.00)"

    autosize_columns(ws, min_width=10, max_width=70)
    return ws
